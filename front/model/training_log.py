"""Training log loader: reads the latest .xlsx from a directory.

Format (per backend spec):
- File: latest .xlsx in the directory (by mtime)
- First row = header (column names are dynamic)
- All values should be numeric
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

from utils.logger import logger


@dataclass
class EpochEntry:
    epoch: int
    metrics: dict[str, float] = field(default_factory=dict)


@dataclass
class TrainingLog:
    status: str                                # "Loaded" | "Missing" | "Error"
    epochs: list[EpochEntry] = field(default_factory=list)
    last_row: dict[str, float] = field(default_factory=dict)
    file_path: Optional[Path] = None
    columns: list[str] = field(default_factory=list)


def load_training_log(log_dir: Path) -> Optional[TrainingLog]:
    if not log_dir.exists() or not log_dir.is_dir():
        logger.info("Training log dir not found: %s", log_dir)
        return TrainingLog(status="Missing")

    xlsx_files = list(log_dir.glob("*.xlsx"))
    if not xlsx_files:
        logger.info("No xlsx in %s", log_dir)
        return TrainingLog(status="Missing")

    latest = max(xlsx_files, key=lambda p: p.stat().st_mtime)

    try:
        wb = load_workbook(latest, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as exc:
        logger.error("Failed to read %s: %s", latest, exc)
        return TrainingLog(status="Error", file_path=latest)

    if not rows:
        return TrainingLog(status="Error", file_path=latest)

    raw_header = rows[0]
    header: list[str] = []
    seen: dict[str, int] = {}
    for i, col in enumerate(raw_header):
        name = str(col).strip() if col is not None else f"col_{i}"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
        header.append(name)

    epochs: list[EpochEntry] = []
    for i, row in enumerate(rows[1:]):
        metrics: dict[str, float] = {}
        for j, val in enumerate(row):
            if j >= len(header):
                continue
            try:
                metrics[header[j]] = float(val) if val is not None else 0.0
            except (TypeError, ValueError):
                pass
        epochs.append(EpochEntry(epoch=i + 1, metrics=metrics))

    last_row = epochs[-1].metrics if epochs else {}

    return TrainingLog(
        status="Loaded",
        epochs=epochs,
        last_row=last_row,
        file_path=latest,
        columns=header,
    )


# ====================================================================
#  Plot column selection (mirrors request.md preferred columns)
# ====================================================================

# Priority order: x-axis column name
_X_CANDIDATES = ("epoch", "Epoch", "step", "iter", "iteration")
# Priority order: primary y (typically loss)
_Y1_CANDIDATES = ("loss", "train_loss", "training_loss")
# Priority order: secondary y (typically acc/ari)
_Y2_CANDIDATES = (
    "ARI", "ari", "acc", "accuracy", "val_acc", "val_accuracy",
    "NMI", "nmi", "HS", "CS",
)


def get_plot_columns(log: TrainingLog) -> tuple[str, Optional[str], Optional[str]]:
    """Return (x_col, y1_col, y2_col) by priority. Falls back gracefully."""
    if not log.epochs:
        return "epoch", None, None

    cols = log.columns

    x_col = next((c for c in _X_CANDIDATES if c in cols), None)
    if not x_col:
        x_col = cols[0]

    y1 = next((c for c in _Y1_CANDIDATES if c in cols), None)
    if not y1:
        for c in cols:
            if c != x_col and c in log.epochs[0].metrics:
                y1 = c
                break

    y2 = next((c for c in _Y2_CANDIDATES if c in cols and c != y1), None)
    if not y2:
        for c in cols:
            if c != x_col and c != y1 and c in log.epochs[0].metrics:
                y2 = c
                break

    return x_col, y1, y2
